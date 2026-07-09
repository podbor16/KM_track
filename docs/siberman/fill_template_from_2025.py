"""
Заполняет пустой чистый шаблон (excel_template_raw.xlsx, скопированный в DST)
реальными данными гонки 2025 года, извлечёнными из уже смигрированного файла
2025_import_ready.xlsx, эмулируя работу судей на новом шаблоне.

Вело (день 1 и день 2) записывается АСТРОНОМИЧЕСКИМ временем — так, как теперь
должны вносить судьи. Плавание/бег остаются elapsed, как раньше.

И старт дня 1 (плавание), и базовый старт дня 2 (лидер по рангу вело-1) —
08:00:00 по красноярскому времени (см. knowledge/decisions/2026-07-09-siberman-live-v2-tasks.md).
Реальный индивидуальный старт вело-2 участника — это уже распарсенное
result.handicaps[cp_key] из исходного файла (бывшая "Стартовая минута").

Плюс добавляет тестового участника (Подборский Матвей) для ручной проверки
после загрузки через /siberman/admin.

Использование:
    python docs/siberman/fill_template_from_2025.py <src 2025_import_ready.xlsx> <dst пустой шаблон.xlsx>
"""
import random
import sys
from typing import Optional
import openpyxl

from src.siberman.parser import (
    parse_excel, parse_time_to_seconds, _normalize_header, _build_col_index, _find_header_row,
)

RACE_START_S = 8 * 3600  # 08:00:00 — старт дня 1 (плавание)

STAGE_SEQ_COUNT = {"swim": 7, "bike_day1": 6, "bike_day2": 8, "run": 12}

# Колонки шаблона (1-based) по каждому этапу/seq — БЕЗ колонки "Стартовая минута"
SWIM_COL = {1: 15, 2: 16, 3: 17, 4: 18, 5: 19, 6: 20, 7: 21}
BIKE1_COL = {1: 22, 2: 23, 3: 24, 4: 25, 5: 26, 6: 27}
BIKE2_COL = {1: 28, 2: 29, 3: 30, 4: 31, 5: 32, 6: 33, 7: 34, 8: 35}
RUN_COL = {i: 35 + i for i in range(1, 13)}  # 36..47


def hms(seconds) -> str:
    if seconds is None:
        return ""
    h, rem = divmod(int(seconds), 3600)
    m, s = divmod(rem, 60)
    return f"{h}:{m:02d}:{s:02d}"


def _write_swim_run(ws, row_i: int, cp: dict) -> None:
    for seq, col in SWIM_COL.items():
        v = cp.get(("swim", seq))
        if v is not None:
            ws.cell(row=row_i, column=col, value=hms(v))
    for seq, col in RUN_COL.items():
        v = cp.get(("run", seq))
        if v is not None:
            ws.cell(row=row_i, column=col, value=hms(v))


def _write_bike_astronomical(ws, row_i: int, cp: dict, bike_day2_start_s: Optional[int]) -> None:
    """bike_day1/bike_day2 в cp сейчас elapsed (как хранит 2025_import_ready.xlsx) —
    записываем в шаблон как astronomical = elapsed + старт соответствующего дня."""
    for seq, col in BIKE1_COL.items():
        v = cp.get(("bike_day1", seq))
        if v is not None:
            ws.cell(row=row_i, column=col, value=hms(RACE_START_S + v))
    if bike_day2_start_s is None:
        return
    for seq, col in BIKE2_COL.items():
        v = cp.get(("bike_day2", seq))
        if v is not None:
            ws.cell(row=row_i, column=col, value=hms(bike_day2_start_s + v))


def _read_legacy_handicaps(src_path: str) -> dict[str, int]:
    """2025_import_ready.xlsx ещё содержит старую колонку "Стартовая минута"
    (parser.py её больше не читает — в новых шаблонах этой колонки нет,
    старт вело-2 теперь считается автоматически). Читаем её напрямую — это
    единственный источник реального старта вело-2 для исторических данных."""
    wb = openpyxl.load_workbook(src_path, data_only=True)
    sheet = wb.active
    header_row_idx = _find_header_row(sheet)
    raw_headers = [c.value for c in sheet[header_row_idx]]
    if header_row_idx > 1:
        prev = [c.value for c in sheet[header_row_idx - 1]]
        raw_headers = [
            h if h is not None else (prev[i] if i < len(prev) else None)
            for i, h in enumerate(raw_headers)
        ]
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    col_idx = _build_col_index(headers)

    handicap_col = next(
        (ci for (n, _), ci in col_idx.items() if n.startswith("стартовая минута")), None
    )
    if handicap_col is None:
        return {}

    fmt_col = col_idx.get(("формат", 0))
    bib_col = col_idx.get(("номер", 0))

    handicaps: dict[str, int] = {}
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue
        bib_raw = row[bib_col] if bib_col is not None else None
        if not bib_raw:
            continue
        bib = str(bib_raw).strip()
        fmt_raw = str(row[fmt_col] or "").strip() if fmt_col is not None else ""
        key = f"{bib}:bike" if fmt_raw in ("Эстафета", "relay") else bib
        if handicap_col < len(row):
            v = parse_time_to_seconds(row[handicap_col])
            if v is not None:
                handicaps[key] = v
    return handicaps


def main(SRC: str, DST: str):
    with open(SRC, "rb") as f:
        data = f.read()
    result = parse_excel(data, 2025)
    result.handicaps = _read_legacy_handicaps(SRC)

    wb = openpyxl.load_workbook(DST)
    ws = wb["Результаты"]

    row_i = 3

    # --- Личники ---
    for p in result.participants:
        if p["format"] != "individual":
            continue
        cp = result.checkpoint_times.get(p["_cp_key"], {})
        bike2_start_s = result.handicaps.get(p["_cp_key"])

        ws.cell(row=row_i, column=1, value="Лично")
        ws.cell(row=row_i, column=2, value=p["bib"])
        ws.cell(row=row_i, column=3, value=p["surname"])
        ws.cell(row=row_i, column=4, value=p["name"])
        ws.cell(row=row_i, column=5, value={"M": "М", "F": "Ж"}.get(p["gender"], p["gender"]))
        ws.cell(row=row_i, column=13, value=p["country"])
        ws.cell(row=row_i, column=14, value=p["city"])
        _write_swim_run(ws, row_i, cp)
        _write_bike_astronomical(ws, row_i, cp, bike2_start_s)
        row_i += 1

    # --- Эстафеты: группируем 3 записи (swim/bike/run) по bib в одну строку ---
    relay_by_bib: dict[str, dict] = {}
    for p in result.participants:
        if p["format"] != "relay":
            continue
        relay_by_bib.setdefault(p["bib"], {})[p["relay_stage"]] = p

    def _full_name(m: Optional[dict]) -> str:
        return f"{m['surname']} {m['name']}".strip() if m else ""

    def _gender_ru(m: Optional[dict]) -> str:
        if not m:
            return ""
        return {"M": "М", "F": "Ж"}.get(m.get("gender"), "")

    for bib, members in relay_by_bib.items():
        any_member = next(iter(members.values()))
        ws.cell(row=row_i, column=1, value="Эстафета")
        ws.cell(row=row_i, column=2, value=bib)
        ws.cell(row=row_i, column=6, value=any_member["relay_team_name"])
        ws.cell(row=row_i, column=7, value=_full_name(members.get("swim")))
        ws.cell(row=row_i, column=8, value=_gender_ru(members.get("swim")))
        ws.cell(row=row_i, column=9, value=_full_name(members.get("bike")))
        ws.cell(row=row_i, column=10, value=_gender_ru(members.get("bike")))
        ws.cell(row=row_i, column=11, value=_full_name(members.get("run")))
        ws.cell(row=row_i, column=12, value=_gender_ru(members.get("run")))
        ws.cell(row=row_i, column=13, value=any_member["country"])
        ws.cell(row=row_i, column=14, value=any_member["city"])

        cp: dict = {}
        for stage_name, cp_key_suffix in (("swim", "swim"), ("bike_day1", "bike"), ("bike_day2", "bike"), ("run", "run")):
            team_cp = result.checkpoint_times.get(f"{bib}:{cp_key_suffix}", {})
            for seq in range(1, STAGE_SEQ_COUNT[stage_name] + 1):
                v = team_cp.get((stage_name, seq))
                if v is not None:
                    cp[(stage_name, seq)] = v

        bike2_start_s = result.handicaps.get(f"{bib}:bike")
        _write_swim_run(ws, row_i, cp)
        _write_bike_astronomical(ws, row_i, cp, bike2_start_s)
        row_i += 1

    # --- Тестовый участник для ручной проверки после загрузки ---
    random.seed(42)
    swim_base = 0
    swim_times = []
    for _ in range(7):
        swim_base += random.randint(280, 340)
        swim_times.append(swim_base)
    # Намеренно медленный вело-1 (не должен попасть в топ-5 и сдвинуть
    # реальные ранги/старты вело-2 остальных участников): суммарно ~8ч.
    bike1_base = swim_base
    bike1_times = []
    for _ in range(6):
        bike1_base += random.randint(4800, 5200)
        bike1_times.append(bike1_base)
    bike2_base = 0
    bike2_times = []
    for _ in range(8):
        bike2_base += random.randint(1800, 3600)
        bike2_times.append(bike2_base)
    run_base = 0
    run_times = []
    for _ in range(12):
        run_base += random.randint(2000, 2400)
        run_times.append(run_base)

    test_bike2_start_s = 8 * 3600 + 6 * 60  # 8:06:00 — условный расчётный старт

    ws.cell(row=row_i, column=1, value="Лично")
    ws.cell(row=row_i, column=2, value=9999)
    ws.cell(row=row_i, column=3, value="Подборский")
    ws.cell(row=row_i, column=4, value="Матвей")
    ws.cell(row=row_i, column=5, value="М")
    ws.cell(row=row_i, column=13, value="Россия")
    ws.cell(row=row_i, column=14, value="Красноярск")
    for seq, v in enumerate(swim_times, start=1):
        ws.cell(row=row_i, column=SWIM_COL[seq], value=hms(v))
    for seq, v in enumerate(bike1_times, start=1):
        ws.cell(row=row_i, column=BIKE1_COL[seq], value=hms(RACE_START_S + v))
    for seq, v in enumerate(bike2_times, start=1):
        ws.cell(row=row_i, column=BIKE2_COL[seq], value=hms(test_bike2_start_s + v))
    for seq, v in enumerate(run_times, start=1):
        ws.cell(row=row_i, column=RUN_COL[seq], value=hms(v))

    wb.save(DST)
    print(f"Saved {row_i - 3 + 1} rows (incl. test participant) to {DST}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
