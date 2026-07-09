"""Генератор чистого Excel-шаблона для судей Siberman — только сырые колонки ввода."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Результаты"

# --- Структура колонок: (row1_категория или None, row2_заголовок) ---
COLUMNS = [
    (None, "Формат"),
    (None, "Номер"),
    (None, "Фамилия"),           # только формат "Лично"
    (None, "Имя"),               # только формат "Лично"
    (None, "Пол"),               # только формат "Лично"
    (None, "Название команды"),      # только формат "Эстафета"
    (None, "Пловец"),                # ФИО, только формат "Эстафета"
    (None, "Пол пловца"),            # только формат "Эстафета"
    (None, "Велосипедист"),          # ФИО, только формат "Эстафета"
    (None, "Пол велосипедиста"),     # только формат "Эстафета"
    (None, "Бегун"),                 # ФИО, только формат "Эстафета"
    (None, "Пол бегуна"),            # только формат "Эстафета"
    (None, "Страна"),            # "Лично" — участника; "Эстафета" — общая на команду
    (None, "Город"),             # "Лично" — участника; "Эстафета" — общий на команду
    ("Плавание", "1,3 км"),
    ("Плавание", "2,6 км"),
    ("Плавание", "3,9 км"),
    ("Плавание", "5,2 км"),
    ("Плавание", "6,5 км"),
    ("Плавание", "7,8 км"),
    ("Плавание", "Финиш 10 км"),
    ("Вело день 1", "3 км"),
    ("Вело день 1", "10 км"),
    ("Вело день 1", "72 км (разворот)"),
    ("Вело день 1", "135 км"),
    ("Вело день 1", "142 км"),
    ("Вело день 1", "Финиш 145 км"),
    ("Вело день 2", "Стартовая минута"),
    ("Вело день 2", "51 км"),
    ("Вело день 2", "82 км"),
    ("Вело день 2", "119 км"),
    ("Вело день 2", "160 км (СШГЭС)"),
    ("Вело день 2", "190 км (Кольцо Саяногорск)"),
    ("Вело день 2", "203 км"),
    ("Вело день 2", "265 км"),
    ("Вело день 2", "Финиш 276 км"),
    ("Бег", "1 круг (7 км)"),
    ("Бег", "2 круга (14 км)"),
    ("Бег", "3 круга (21 км)"),
    ("Бег", "4 круга (28 км)"),
    ("Бег", "5 кругов (35 км)"),
    ("Бег", "6 кругов (42 км)"),
    ("Бег", "7 кругов (49 км)"),
    ("Бег", "8 кругов (56 км)"),
    ("Бег", "9 кругов (63 км)"),
    ("Бег", "10 кругов (70 км)"),
    ("Бег", "11 кругов (77 км)"),
    ("Бег", "12 кругов (84 км) - Финиш"),
]

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
cat_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
cat_font = Font(bold=True)

for col_i, (cat, label) in enumerate(COLUMNS, start=1):
    letter = get_column_letter(col_i)
    if cat:
        ws.cell(row=1, column=col_i, value=cat)
        ws.cell(row=1, column=col_i).fill = cat_fill
        ws.cell(row=1, column=col_i).font = cat_font
    c2 = ws.cell(row=2, column=col_i, value=label)
    c2.fill = header_fill
    c2.font = header_font
    c2.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[letter].width = max(12, len(label) * 0.9)

# Объединить категории вело день1/день2/плавание/бег по соседним ячейкам row1
from itertools import groupby
col = 1
for cat, group in groupby(COLUMNS, key=lambda x: x[0]):
    n = len(list(group))
    if cat and n > 1:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n - 1)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    col += n

ws.freeze_panes = "A3"

wb.save(r"C:\Users\podbo\Работа\КРАСМАРАФОН\KM_track\docs\siberman\excel_template_raw.xlsx")
print("Saved.")
