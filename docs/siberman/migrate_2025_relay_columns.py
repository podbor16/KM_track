"""
Разовый конвертер: приводит реальный файл "Live Siberman 2025.xlsx" (старая
схема эстафетных колонок: Фамилия=команда, Имя/Страна/Город=ФИО трёх
участников) к новой схеме, которую понимает src/siberman/parser.py
(Название команды / Пловец / Велосипедист / Бегун).

Страна/Город команды в старом файле не хранились отдельно (эти колонки были
заняты под ФИО) — при конвертации оставляем их пустыми, это ожидаемо.

Использование:
    python docs/siberman/migrate_2025_relay_columns.py <входной .xlsx> <выходной .xlsx>
"""
import sys
import openpyxl


def migrate(src_path: str, dst_path: str, sheet_name: str = "2025") -> None:
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb[sheet_name]

    headers = {str(c.value).strip().lower(): c.column for c in ws[2] if c.value}
    fmt_col = headers["формат"]
    surname_col = headers["фамилия"]
    name_col = headers["имя"]
    country_col = headers["страна"]
    city_col = headers["город"]

    # Вставляем 4 новые колонки сразу после "Город"
    insert_at = city_col + 1
    ws.insert_cols(insert_at, amount=4)
    new_headers = ["Название команды", "Пловец", "Велосипедист", "Бегун"]
    for i, h in enumerate(new_headers):
        ws.cell(row=2, column=insert_at + i, value=h)

    # После вставки колонок индексы старых полей сдвигаются, если они были
    # правее insert_at — но surname/name/country/city все левее (< insert_at),
    # так что их индексы не меняются.
    for row in ws.iter_rows(min_row=3):
        fmt_val = row[fmt_col - 1].value
        if fmt_val != "Эстафета":
            continue
        team_name = row[surname_col - 1].value
        swim_name = row[name_col - 1].value
        bike_name = row[country_col - 1].value
        run_name = row[city_col - 1].value

        ws.cell(row=row[0].row, column=insert_at + 0, value=team_name)
        ws.cell(row=row[0].row, column=insert_at + 1, value=swim_name)
        ws.cell(row=row[0].row, column=insert_at + 2, value=bike_name)
        ws.cell(row=row[0].row, column=insert_at + 3, value=run_name)

        # Очищаем поля, которые в новой схеме относятся только к личникам.
        # ws.cell(..., value=None) — не работает: None здесь означает "значение
        # не передано" (сигнатура openpyxl), а не "очистить". Нужно .value = None.
        row[surname_col - 1].value = None
        row[name_col - 1].value = None
        row[country_col - 1].value = None
        row[city_col - 1].value = None

    wb.save(dst_path)
    print(f"Saved: {dst_path}")


if __name__ == "__main__":
    src, dst = sys.argv[1], sys.argv[2]
    migrate(src, dst)
