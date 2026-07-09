"""
Excel-парсер для данных Siberman.

Формат файла: Google Sheets → Скачать как .xlsx
Ряд 1 или 2 — заголовки (зависит от листа).
Времена: строка "Ч:ММ:СС", timedelta, float (доля суток) или пусто/DNF.

Парсер НЕ различает elapsed и астрономическое время — это просто число
секунд, конвертированное как есть. Для вело (bike_day1/bike_day2) судьи
вносят астрономическое время; перевод в elapsed и расчёт старта вело-2
происходят отдельным шагом в service.py:convert_bike_times_to_elapsed(),
не здесь.

Если заголовок в реальном файле отличается от CHECKPOINT_COL_MAP —
добавь его в маппинг или измени нормализацию в _normalize_header().
"""
import datetime
import io
from dataclasses import dataclass, field
from typing import Optional

import openpyxl


# ---------------------------------------------------------------------------
# Маппинг: (нормализованный заголовок, порядковый индекс дубля) → (этап, seq)
# ---------------------------------------------------------------------------
CHECKPOINT_COL_MAP: dict[tuple[str, int], tuple[str, int]] = {
    # Плавание: заголовки — дистанции отрезков (Row2)
    ("1,3 км", 0):          ("swim", 1),
    ("2,6 км", 0):          ("swim", 2),
    ("3,9 км", 0):          ("swim", 3),
    ("5,2 км", 0):          ("swim", 4),
    ("6,5 км", 0):          ("swim", 5),
    ("7,8 км", 0):          ("swim", 6),
    ("финиш 10 км", 0):     ("swim", 7),
    # Вело день 1
    ("3 км", 0):            ("bike_day1", 1),
    ("10 км", 0):           ("bike_day1", 2),
    ("72 км", 0):           ("bike_day1", 3),   # "72 км (разворот)" — startswith
    ("135 км", 0):          ("bike_day1", 4),
    ("142 км", 0):          ("bike_day1", 5),
    ("финиш 145 км", 0):    ("bike_day1", 6),
    # Вело день 2
    ("51 км", 0):           ("bike_day2", 1),
    ("82 км", 0):           ("bike_day2", 2),
    ("119 км", 0):          ("bike_day2", 3),
    ("160 км", 0):          ("bike_day2", 4),   # "160 км (СШГЭС)" — startswith
    ("190 км", 0):          ("bike_day2", 5),   # "190 км Кольцо Саяногорск" — startswith
    ("203 км", 0):          ("bike_day2", 6),
    ("265 км", 0):          ("bike_day2", 7),
    ("финиш 276 км", 0):    ("bike_day2", 8),
    # Бег: "1 круг (7 км)", "2 круга (14 км)", "5 кругов (35 км)"... — startswith
    ("1 круг", 0):          ("run", 1),
    ("2 круга", 0):         ("run", 2),
    ("3 круга", 0):         ("run", 3),
    ("4 круга", 0):         ("run", 4),
    ("5 кругов", 0):        ("run", 5),
    ("6 кругов", 0):        ("run", 6),
    ("7 кругов", 0):        ("run", 7),
    ("8 кругов", 0):        ("run", 8),
    ("9 кругов", 0):        ("run", 9),
    ("10 кругов", 0):       ("run", 10),
    ("11 кругов", 0):       ("run", 11),
    ("12 кругов", 0):       ("run", 12),
}

# Колонки с данными участника: нормализованный заголовок → поле в dict
PARTICIPANT_COL_MAP: dict[str, str] = {
    "формат":  "format",
    "номер":   "bib",
    "фамилия": "surname",
    "имя":     "name",
    "страна":  "country",
    "город":   "city",
    "пол":     "gender",   # М/Ж — добавить в Excel если отсутствует
    # Только для строк формата "Эстафета" — страна/город выше используются как общие для команды
    "название команды":     "relay_team_name",
    "пловец":               "relay_swim_name",
    "велосипедист":         "relay_bike_name",
    "бегун":                "relay_run_name",
    "пол пловца":           "relay_swim_gender",
    "пол велосипедиста":    "relay_bike_gender",
    "пол бегуна":           "relay_run_gender",
}

# Специальные колонки
T1_COL = "t1"
T2_COL = "t2"


@dataclass
class ParseResult:
    race_year: int
    participants: list[dict] = field(default_factory=list)
    # {bib: {(stage, seq): cumulative_s}} — для bike_day1/bike_day2 хранит
    # RAW астрономическое время (секунды от полуночи) до конвертации через
    # service.py:convert_bike_times_to_elapsed()
    checkpoint_times: dict[str, dict[tuple[str, int], Optional[int]]] = field(default_factory=dict)
    # {bib: {zone: duration_s}}
    transitions: dict[str, dict[str, Optional[int]]] = field(default_factory=dict)
    # {bib_or_relay_bike_key: start_day2_s} — заполняется НЕ парсером, а
    # convert_bike_times_to_elapsed() в service.py (расчётный старт вело-2)
    handicaps: dict[str, Optional[int]] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


def parse_time_to_seconds(value) -> Optional[int]:
    """Конвертировать время в секунды. None — пусто или DNF."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if not v or v.upper() in ("DNF", "DNS", "DSQ", "ДНФ", "-", "—"):
            return None
        # Handle lowercase cyrillic DNF variants
        if v.lower() in ("днф", "днс", "дсq"):
            return None
        parts = v.split(":")
        try:
            if len(parts) == 3:
                return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
            if len(parts) == 2:
                return int(parts[0]) * 60 + int(parts[1])
        except ValueError:
            return None
        return None
    if isinstance(value, datetime.timedelta):
        return int(value.total_seconds())
    if isinstance(value, datetime.time):
        return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, (int, float)):
        return int(round(value * 86400))
    return None


def _classify_time_cell(value) -> tuple[Optional[int], Optional[str]]:
    """Как parse_time_to_seconds, но также возвращает причину, если None
    означает "не удалось распознать" (а не легитимно пустое/DNF)."""
    seconds = parse_time_to_seconds(value)
    if seconds is not None or value is None:
        return seconds, None
    if isinstance(value, str):
        v = value.strip()
        if not v or v.upper() in ("DNF", "DNS", "DSQ", "ДНФ", "-", "—") or v.lower() in ("днф", "днс", "дсq"):
            return None, None
        return None, f"не удалось распознать время «{value}» (ожидался формат Ч:ММ:СС)"
    return None, f"неожиданный тип значения времени: {value!r}"


def _normalize_header(h: str) -> str:
    return str(h).strip().lower()


def _normalize_gender(raw: str) -> Optional[str]:
    g = raw.strip().upper()
    if g in ("Ж", "F", "FEMALE"):
        return "F"
    if g in ("М", "M", "MALE"):
        return "M"
    return None


def _build_col_index(headers: list[str]) -> dict[tuple[str, int], int]:
    """Построить маппинг (normalized, occurrence) → col_index."""
    counts: dict[str, int] = {}
    result: dict[tuple[str, int], int] = {}
    for idx, h in enumerate(headers):
        n = _normalize_header(h)
        occ = counts.get(n, 0)
        counts[n] = occ + 1
        result[(n, occ)] = idx
    return result


def _find_header_row(sheet) -> int:
    """Вернуть индекс строки с заголовками (1-based). Ищет строку с 'Номер' или 'номер'."""
    for i, row in enumerate(sheet.iter_rows(max_row=5, values_only=True), start=1):
        for cell in row:
            if cell and str(cell).strip().lower() in ("номер", "bib"):
                return i
    return 1  # fallback


def parse_excel(file_bytes: bytes, race_year: int) -> ParseResult:
    """Разобрать Excel-файл формата 2025. Возвращает ParseResult."""
    result = ParseResult(race_year=race_year)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    header_row_idx = _find_header_row(sheet)
    raw_headers = [cell.value for cell in sheet[header_row_idx]]
    # Если заголовки в двух строках (Row1 — категории, Row2 — подзаголовки),
    # подставляем значение из Row1 там, где Row2 пуста (напр. "Стартовая минута").
    if header_row_idx > 1:
        prev_headers = [cell.value for cell in sheet[header_row_idx - 1]]
        raw_headers = [
            h if h is not None else (prev_headers[i] if i < len(prev_headers) else None)
            for i, h in enumerate(raw_headers)
        ]
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    col_idx = _build_col_index(headers)

    # Инвертированный маппинг col_index → (stage, seq) для чекпоинтов
    cp_by_col: dict[int, tuple[str, int]] = {}
    for (norm, occ), stage_seq in CHECKPOINT_COL_MAP.items():
        for (n, o), ci in col_idx.items():
            if o == occ and (n == norm or n.startswith(norm + " ") or n.startswith(norm + "/")):
                cp_by_col[ci] = stage_seq
                break

    # Индексы колонок участника
    part_col: dict[str, int] = {}
    for norm_name, field_name in PARTICIPANT_COL_MAP.items():
        for (n, _), ci in col_idx.items():
            if n == norm_name:
                part_col[field_name] = ci
                break

    # Индекс T1
    t1_col = col_idx.get(("t1", 0))

    # Парсить строки данных. Не values_only, чтобы иметь адрес ячейки (для errors).
    for row_cells in sheet.iter_rows(min_row=header_row_idx + 1):
        row = [c.value for c in row_cells]
        if not any(row):
            continue

        bib_col = part_col.get("bib")
        bib_raw = row[bib_col] if bib_col is not None else None
        if not bib_raw:
            first_ci = next((i for i, v in enumerate(row) if v not in (None, "")), 0)
            coord = row_cells[first_ci].coordinate
            result.errors.append(f"{coord}: в строке есть данные, но не заполнен номер участника — строка пропущена")
            continue
        bib = str(bib_raw).strip()

        def _cell(f: str, default: str = "", _row=row) -> str:
            ci = part_col.get(f)
            return str(_row[ci] or default).strip() if ci is not None and ci < len(_row) else default

        fmt_raw = _cell("format", "Лично")
        fmt = "relay" if fmt_raw in ("Эстафета", "relay") else "individual"

        def _check_gender(field: str, role_label: str) -> None:
            raw = _cell(field)
            if raw and _normalize_gender(raw) is None:
                ci = part_col.get(field)
                coord = row_cells[ci].coordinate if ci is not None else f"строка {row_cells[0].row}"
                result.errors.append(
                    f"{coord} (участник {bib}, «{role_label}»): пол «{raw}» не распознан, использовано значение по умолчанию «М»"
                )

        # Чекпоинты — общие для строки
        cp_times: dict[tuple[str, int], Optional[int]] = {}
        for ci, stage_seq in cp_by_col.items():
            if ci < len(row):
                seconds, err = _classify_time_cell(row[ci])
                cp_times[stage_seq] = seconds
                if err:
                    coord = row_cells[ci].coordinate
                    label = headers[ci] if ci < len(headers) else f"колонка {ci + 1}"
                    result.errors.append(f"{coord} (участник {bib}, «{label}»): {err}")

        if fmt == "relay":
            # Название команды + ФИО участников swim/bike/run — отдельные колонки.
            # Страна/Город — общие на команду (не теряются, в отличие от старой схемы).
            team_name = _cell("relay_team_name")
            members_raw = [_cell("relay_swim_name"), _cell("relay_bike_name"), _cell("relay_run_name")]
            members_gender_raw = [_cell("relay_swim_gender"), _cell("relay_bike_gender"), _cell("relay_run_gender")]
            team_country = _cell("country", "Россия")
            team_city = _cell("city")
            relay_stages = ["swim", "bike", "run"]

            def _split_name(full: str) -> tuple[str, str]:
                parts = full.strip().split(None, 1)
                return (parts[0], parts[1]) if len(parts) >= 2 else (full.strip(), "")

            base = {
                "race_year": race_year, "bib": bib, "format": "relay",
                "relay_team_name": team_name,
                "country": team_country, "city": team_city, "status": "active",
            }
            relay_gender_fields = {"swim": "relay_swim_gender", "bike": "relay_bike_gender", "run": "relay_run_gender"}
            relay_role_labels = {"swim": "пол пловца", "bike": "пол велосипедиста", "run": "пол бегуна"}
            for rs, full_name, gender_raw in zip(relay_stages, members_raw, members_gender_raw):
                sn, nm = _split_name(full_name)
                _check_gender(relay_gender_fields[rs], relay_role_labels[rs])
                result.participants.append({
                    **base,
                    "surname": sn, "name": nm, "relay_stage": rs,
                    "gender": _normalize_gender(gender_raw) or "M",
                    "_cp_key": f"{bib}:{rs}",
                })

            # Разбить чекпоинты по этапу участника
            result.checkpoint_times[f"{bib}:swim"] = {
                k: v for k, v in cp_times.items() if k[0] == "swim"
            }
            result.checkpoint_times[f"{bib}:bike"] = {
                k: v for k, v in cp_times.items() if k[0] in ("bike_day1", "bike_day2")
            }
            result.checkpoint_times[f"{bib}:run"] = {
                k: v for k, v in cp_times.items() if k[0] == "run"
            }

        else:
            _check_gender("gender", "пол")
            gender = _normalize_gender(_cell("gender", "")) or "M"

            has_dnf = any(
                cell and str(cell).strip().upper() in ("DNF", "ДНФ") for cell in row
            )
            result.participants.append({
                "race_year": race_year, "bib": bib,
                "surname":   _cell("surname"), "name": _cell("name"),
                "gender":    gender,
                "country":   _cell("country", "Россия"), "city": _cell("city"),
                "format":    "individual",
                "status":    "dnf" if has_dnf else "active",
                "_cp_key":   bib,
            })
            result.checkpoint_times[bib] = cp_times

            if t1_col is not None and t1_col < len(row):
                t1 = parse_time_to_seconds(row[t1_col])
                if t1 is not None:
                    result.transitions[bib] = {"T1": t1}

    return result
